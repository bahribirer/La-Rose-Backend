from typing import List, Dict
import openpyxl
from io import BytesIO
from app.core.database import db
from app.pharmacies.utils import normalize_text
import logging

logger = logging.getLogger(__name__)

async def process_pharmacy_excel(file_content: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    pharmacies = []
    
    logger.info(f"Sheets found: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        league = sheet_name.strip()
        
        # Determine column indices by searching for headers in the first 5 rows
        header_row_idx = -1
        col_map = {}
        
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            row_vals = [str(cell).upper() if cell else "" for cell in row]
            if "ECZANE ADI" in row_vals or "MÜMESSİL" in "".join(row_vals) or "MUMESSIL" in "".join(row_vals):
                header_row_idx = i
                for idx, val in enumerate(row_vals):
                    if "MÜMESSİL" in val or "MUMESSIL" in val: col_map["representative"] = idx
                    elif "ECZANE ADI" in val: col_map["pharmacy_name"] = idx
                    elif "ECZACI" in val: col_map["pharmacist"] = idx
                    elif "ŞEHİR" in val or "BÖLGE" in val: col_map["district"] = idx
                break
        
        # Fallback to fixed indexing if headers not found
        if "pharmacy_name" not in col_map: 
            col_map = {"representative": 1, "pharmacy_name": 3, "pharmacist": 4, "district": 5}
            header_row_idx = 1
            
        logger.info(f"Processing sheet {league} with mapping {col_map} starting from row {header_row_idx + 1}")

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or len(row) < 4: continue
            
            name = str(row[col_map["pharmacy_name"]]).strip() if row[col_map["pharmacy_name"]] else None
            if not name or name.lower() in ["none", "eczane adi", "null", ""]: continue
            
            rep = str(row[col_map["representative"]]).strip() if row[col_map["representative"]] else "Atanmamış"
            dist = str(row[col_map["district"]]).strip() if len(row) > col_map["district"] and row[col_map["district"]] else "-"
            pharmacist = str(row[col_map["pharmacist"]]).strip() if len(row) > col_map["pharmacist"] and row[col_map["pharmacist"]] else "-"

            pharmacies.append({
                "pharmacy_name": name,
                "normalized_name": normalize_text(name),
                "league": league,
                "representative": rep,
                "district": dist,
                "pharmacist": pharmacist
            })

    if not pharmacies:
        logger.warning("No pharmacies found in any sheet.")
        return 0

    await db.pharmacies.delete_many({})
    await db.pharmacies.insert_many(pharmacies)
    logger.info(f"Successfully imported {len(pharmacies)} pharmacies.")
    return len(pharmacies)

async def get_pharmacies_list(league: str = None, representative: str = None):
    query = {}
    if league:
        query["league"] = league
    if representative:
        query["representative"] = representative
        
    cursor = db.pharmacies.find(query).sort("league", 1)
    results = []
    async for doc in cursor:
        rep = doc.get("representative")
        # Handle cases where representative might be an object in old data
        if isinstance(rep, dict):
            rep = rep.get("name") or str(rep)

        results.append({
            "id": str(doc["_id"]),
            "pharmacy_name": doc["pharmacy_name"],
            "league": doc.get("league") or "Diğer",
            "representative": rep,
            "district": doc.get("district"),
            "pharmacist": doc.get("pharmacist"),
        })
    return results

async def create_pharmacy(data: dict):
    # Normalize name if present
    if "pharmacy_name" in data:
        data["normalized_name"] = normalize_text(data["pharmacy_name"])
    
    result = await db.pharmacies.insert_one(data)
    data["id"] = str(result.inserted_id)
    return data

async def update_pharmacy(pharmacy_id: str, data: dict):
    # Normalize name if present
    if "pharmacy_name" in data:
        data["normalized_name"] = normalize_text(data["pharmacy_name"])
    
    from bson import ObjectId
    try:
        obj_id = ObjectId(pharmacy_id)
    except:
        return None
        
    result = await db.pharmacies.update_one(
        {"_id": obj_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        return None
        
    return await db.pharmacies.find_one({"_id": obj_id})

async def delete_pharmacy(pharmacy_id: str):
    from bson import ObjectId
    try:
        obj_id = ObjectId(pharmacy_id)
    except:
        return False
        
    result = await db.pharmacies.delete_one({"_id": obj_id})
    return result.deleted_count > 0
