from typing import Optional
from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from datetime import datetime
from app.admin.dependencies import admin_required
from app.competitions.utils import end_of_month_utc
from app.core.database import db
from app.core.utils import serialize_mongo


from app.admin.service import (
    get_overview,
    get_top_users,
    get_top_products
) 
from app.admin.schemas import AdminOverviewResponse

router = APIRouter(
    prefix="/admin",
    tags=["Admin"]
)


@router.get("/overview", response_model=AdminOverviewResponse)
async def admin_overview(
    start: datetime = Query(..., description="UTC datetime (ISO 8601)"),
    end: datetime = Query(..., description="UTC datetime (ISO 8601)"),
    _=Depends(admin_required),
):
    return {
        "overview": await get_overview(start, end),
        "top_users": await get_top_users(start, end),
        "top_products": await get_top_products(start, end, limit=10),
    }


@router.get("/products", dependencies=[Depends(admin_required)])
async def admin_products_list(
    start: datetime = Query(..., description="UTC datetime (ISO 8601)"),
    end: datetime = Query(..., description="UTC datetime (ISO 8601)"),
):
    # Ürün detay sayfasında tüm ürünleri görmek için limit=None
    return await get_top_products(start, end, limit=None)


@router.get("/analytics/daily", dependencies=[Depends(admin_required)])
async def daily_analytics(
    start: datetime = Query(..., description="UTC datetime (ISO 8601)"),
    end: datetime = Query(..., description="UTC datetime (ISO 8601)"),
):
    pipeline = [
        {
            "$match": {
                "createdAt": {
                    "$gte": start,
                    "$lte": end
                }
            }
        },
        {
            "$group": {
                "_id": {
                    "$dateToString": {
                        "format": "%Y-%m-%d",
                        "date": "$createdAt"
                    }
                },
                "total_profit": {
                    "$sum": "$summary.total_profit"
                },
                "total_cost": {
                    "$sum": "$summary.total_cost"
                }
            }
        },
        {
            "$addFields": {
                "total_revenue": {
                    "$add": ["$total_profit", "$total_cost"]
                }
            }
        },
        {"$sort": {"_id": 1}}
    ]

    cursor = db.sales_reports.aggregate(pipeline)

    labels, revenue, profit = [], [], []
    async for r in cursor:
        labels.append(r["_id"])
        revenue.append(float(r["total_revenue"]))
        profit.append(float(r["total_profit"]))

    return {
        "labels": labels,
        "revenue": revenue,
        "profit": profit
    }
@router.get("/users", dependencies=[Depends(admin_required)])
async def list_users():
    pipeline = [
        # 🔹 PROFILE JOIN
        {
            "$lookup": {
                "from": "user_profiles",
                "localField": "_id",
                "foreignField": "user_id",
                "as": "profile"
            }
        },
        {
            "$unwind": {
                "path": "$profile",
                "preserveNullAndEmptyArrays": True
            }
        },

        # 🔥 PROFILE FLATTEN
        {
  "$addFields": {
    "pharmacy_name": "$profile.pharmacy_name",
    "district": "$profile.district",
    "region": "$profile.region",

    # 🔥 REPRESENTATIVE NORMALIZE
    "representative": {
      "$cond": {
        "if": { "$eq": [{ "$type": "$profile.representative" }, "object"] },
        "then": "$profile.representative.name",
        "else": "$profile.representative"
      }
    }
  }
},


        # 🔹 SALES REPORTS JOIN
        {
            "$lookup": {
                "from": "sales_reports",
                "localField": "_id",
                "foreignField": "user_id",
                "as": "reports"
            }
        },

        # 🔹 AGGREGATES
        {
            "$addFields": {
                "total_reports": { "$size": "$reports" },
                "total_profit": { "$sum": "$reports.summary.total_profit" },
                "total_cost": { "$sum": "$reports.summary.total_cost" },
                "total_items": { "$sum": "$reports.summary.total_items" }
            }
        },

        # 🔹 PROJECT (⚠️ SADECE FIELD: 1)
        {
            "$project": {
                "id": { "$toString": "$_id" },
                "email": 1,
                "full_name": 1,
                "avatar": 1,
                "role": 1,
                "created_at": 1,

                "pharmacy_name": 1,
                "district": 1,
                "region": 1,
                "representative": 1,

                "total_reports": 1,
                "total_profit": 1,
                "total_cost": 1,
                "total_items": 1
            }
        },

        # 🔹 SORT
        {
            "$sort": { "created_at": -1 }
        }
    ]

    cursor = db.users.aggregate(pipeline)

    users = []
    async for u in cursor:
        users.append({
            "id": str(u["_id"]),
            "email": u.get("email"),
            "full_name": u.get("full_name"),
            "avatar": u.get("avatar"),
            "role": u.get("role", "user"),
            "created_at": u.get("created_at"),

            "pharmacy_name": u.get("pharmacy_name"),
            "district": u.get("district"),
            "region": u.get("region"),
            "representative": u.get("representative"),

            "total_reports": int(u.get("total_reports", 0)),
            "total_profit": float(u.get("total_profit", 0)),
            "total_revenue": float(u.get("total_profit", 0) + u.get("total_cost", 0)),
            "total_items": int(u.get("total_items", 0)),
        })

    return users



@router.get("/reports", dependencies=[Depends(admin_required)])
async def admin_reports():
    cursor = (
        db.sales_reports
        .find({})
        .sort("createdAt", -1)
    )

    # ================== 1️⃣ PREPARE DATA ==================
    # Tüm yarışmaları çek (Tarih eşleşmesi için)
    competitions = await db.competitions.find().to_list(None)
    
    # Kullanıcıları bul (İsim göstermek için)
    user_ids = await db.sales_reports.distinct("user_id")
    users = await db.users.find(
        {"_id": {"$in": user_ids}},
        {"_id": 1, "full_name": 1, "email": 1}
    ).to_list(None)
    
    user_map = {
        u["_id"]: (u.get("full_name") or u.get("email") or "Kullanıcı") 
        for u in users
    }

    # ================== 2️⃣ PROCESS REPORTS ==================
    items = []
    async for r in cursor:
        summary = r.get("summary", {
            "total_items": 0,
            "total_profit": 0,
            "total_cost": 0,
        })
        
        if "total_revenue" not in summary:
            summary["total_revenue"] = summary.get("total_sales") or (summary.get("total_profit", 0) + summary.get("total_cost", 0))

        # 📂 FOLDER MAPPING
        folder = None
        
        # 1. Öncelik: Doğrudan ID
        cid = r.get("competition_id")
        if cid:
            # ID üzerinden bul
            comp = next((c for c in competitions if c["_id"] == cid), None)
            if comp:
                folder = {"year": comp["year"], "month": comp["month"]}
        
        # 2. Öncelik: Tarih Aralığı (ID yoksa)
        if not folder and "createdAt" in r:
            created = r["createdAt"]
            # Hangi yarışmanın aralığında?
            for c in competitions:
                # Sadece active/completed yarışmalara bak
                if c.get("status") not in ["active", "completed"]:
                    continue
                    
                if c.get("starts_at") and c.get("ends_at"):
                     if c["starts_at"] <= created <= c["ends_at"]:
                         folder = {"year": c["year"], "month": c["month"]}
                         break
        
        # 👤 USER NAME
        user_name = user_map.get(r.get("user_id"), "Bilinmeyen Kullanıcı")

        items.append({
            "id": str(r["_id"]),
            "name": r.get("name"),
            "user_name": user_name, # 🔥 Yeni Kolon
            "createdAt": r.get("createdAt"),
            "summary": summary,
            "is_competition_report": bool(r.get("is_competition_report", False)),
            "folder": folder, # 🔥 Akıllı Gruplama
        })

    return items

@router.get("/reports/{report_id}", dependencies=[Depends(admin_required)])
async def admin_report_detail(report_id: str):
    if not ObjectId.is_valid(report_id):
        raise HTTPException(status_code=400, detail="Geçersiz rapor ID")

    report = await db.sales_reports.find_one(
        {"_id": ObjectId(report_id)}
    )

    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")

    cursor = db.sales_items.find(
        {"report_id": ObjectId(report_id)}
    )

    # 🔍 Fetch all products once for lookup
    all_products = await db.products.find({}).to_list(None)
    prod_meta_map = {p["id"]: p for p in all_products}

    items = []
    async for i in cursor:
        p_id = i.get("productId")
        prod_meta = prod_meta_map.get(p_id, {})
        barcode = p_id if p_id else "-"
        
        # Prefer tr_name > updated name > historical scan name
        display_name = prod_meta.get("tr_name") or prod_meta.get("name") or i.get("productName")

        items.append({
            "id": str(i["_id"]),
            "product_id": p_id,
            "product_name": display_name,
            "quantity": int(i.get("quantity") or 0),
            "unit_price": float(i.get("unitPrice") or i.get("unit_price") or 0),
            "total_price": float(i.get("totalPrice") or i.get("total_price") or 0),
            "unitPrice": float(i.get("unitPrice") or i.get("unit_price") or i.get("birim_fiyat") or 0),
            "totalPrice": float(i.get("totalPrice") or i.get("total_price") or i.get("tutar") or 0),
            "price": float(i.get("unitPrice") or i.get("unit_price") or i.get("birim_fiyat") or 0),
            "profit": float(i.get("profit") or i.get("ecz_kar") or 0),
            "cost": float(i.get("cost") or i.get("maliyet") or 0),
            "confidence": float(i.get("confidence") or i.get("match_confidence") or 0),
            "date": i.get("date"),
            "barcode": barcode,
            "stock": int(i.get("stock") or i.get("stok_miktari") or 0)
        })

    # 🔥 FIX: Admin Panel needs total_revenue (profit + cost)
    summary = report.get("summary", {})
    if "total_revenue" not in summary:
        summary["total_revenue"] = summary.get("total_sales") or (summary.get("total_profit", 0) + summary.get("total_cost", 0) + 0.0001)

    print(f"🚀 ADMIN DETAIL RESP: {summary} | ITEM[0]: {items[0] if items else 'EMPTY'}")

    return {
        "id": str(report["_id"]),
        "name": report.get("name", "Satış Raporu"),
        "createdAt": report.get("createdAt"),
        "summary": summary,
        "items": items
    }

# 🔥 DEBUG ENDPOINT: PROOF OF DATA
@router.get("/debug/report-item/{report_id}", dependencies=[Depends(admin_required)])
async def debug_report_item(report_id: str):
    item = await db.sales_items.find_one({"report_id": ObjectId(report_id)})
    if item:
        item["_id"] = str(item["_id"])
        item["report_id"] = str(item["report_id"])
        return item
    return {"error": "No items found for this report"}

@router.get("/representatives", dependencies=[Depends(admin_required)])
async def representatives_performance(
    region: Optional[str] = Query(default=None),
):
    pipeline = []

    # 🔥 BÖLGE FİLTRESİ (OPSİYONEL)
    if region:
        pipeline.append({
            "$match": { "region": region }
        })

    pipeline += [
        # sadece eşleşmiş profiller
        {
            "$match": {
                "representative": { "$ne": None }
            }
        },

        # 🔹 normalize (object gelirse stringe çevir)
        {
            "$addFields": {
                "rep_name": {
                    "$cond": {
                        "if": { "$eq": [{ "$type": "$representative" }, "object"] },
                        "then": "$representative.name",
                        "else": "$representative"
                    }
                }
            }
        },

        # 🔹 group by representative + region
        {
            "$group": {
                "_id": {
                    "representative": "$rep_name",
                    "region": "$region"
                },
                "pharmacy_ids": { "$addToSet": "$pharmacy_id" },
                "user_ids": { "$addToSet": "$user_id" }
            }
        },

        # 🔹 counts
        {
            "$addFields": {
                "pharmacy_count": { "$size": "$pharmacy_ids" },
                "user_count": { "$size": "$user_ids" }
            }
        },

        # 🔹 join sales_reports
        {
            "$lookup": {
                "from": "sales_reports",
                "localField": "user_ids",
                "foreignField": "user_id",
                "as": "reports"
            }
        },

        # 🔹 aggregates
        {
            "$addFields": {
                "total_items": { "$sum": "$reports.summary.total_items" },
                "total_profit": { "$sum": "$reports.summary.total_profit" },
                "total_cost": { "$sum": "$reports.summary.total_cost" }
            }
        },
        { 
            "$addFields": { 
                "total_revenue": { "$add": ["$total_profit", "$total_cost"] } 
            } 
        },

        # 🔹 sıralama
        { "$sort": { "total_revenue": -1 } }
    ]

    cursor = db.user_profiles.aggregate(pipeline)

    results = []
    async for r in cursor:
        results.append({
            "representative": r["_id"]["representative"],
            "region": r["_id"]["region"],
            "pharmacy_count": int(r.get("pharmacy_count", 0)),
            "user_count": int(r.get("user_count", 0)),
            "total_items": int(r.get("total_items", 0)),
            "total_profit": float(r.get("total_profit", 0)),
            "total_revenue": float(r.get("total_revenue", 0)),
        })

    return results

from fastapi import Depends, HTTPException
from urllib.parse import unquote

@router.get("/representatives/{name}", dependencies=[Depends(admin_required)])
async def representative_detail(name: str):
    rep_name = unquote(name)

    pipeline = [
        # ilgili mümessil
        { "$match": { "representative": { "$ne": None } } },

        # normalize (object → string)
        {
            "$addFields": {
                "rep_name": {
                    "$cond": {
                        "if": { "$eq": [{ "$type": "$representative" }, "object"] },
                        "then": "$representative.name",
                        "else": "$representative"
                    }
                }
            }
        },
        { "$match": { "rep_name": rep_name } },

        # kullanıcı & eczane setleri
        {
            "$group": {
                "_id": {
                    "representative": "$rep_name",
                    "region": "$region"
                },
                "user_ids": { "$addToSet": "$user_id" },
                "pharmacies": { "$addToSet": "$pharmacy_name" }
            }
        },

        # sales join
        {
            "$lookup": {
                "from": "sales_reports",
                "localField": "user_ids",
                "foreignField": "user_id",
                "as": "reports"
            }
        },

        # totals
        {
            "$addFields": {
                "total_items": { "$sum": "$reports.summary.total_items" },
                "total_profit": { "$sum": "$reports.summary.total_profit" },
                "total_cost": { "$sum": "$reports.summary.total_cost" }
            }
        }
    ]

    cursor = db.user_profiles.aggregate(pipeline)
    doc = await cursor.to_list(length=1)

    if not doc:
        raise HTTPException(404, "Mümessil bulunamadı")

    r = doc[0]

    # kullanıcı listesi (isim + email)
    users = []
    async for u in db.users.find({ "_id": { "$in": r["user_ids"] } }):
        users.append({
            "id": str(u["_id"]),
            "name": u.get("full_name") or u.get("email"),
            "email": u.get("email")
        })

    return {
        "representative": r["_id"]["representative"],
        "region": r["_id"]["region"],
        "pharmacies": sorted([p for p in r.get("pharmacies", []) if p]),
        "users": users,
        "total_items": int(r.get("total_items", 0)),
        "total_profit": float(r.get("total_profit", 0)),
        "total_revenue": float(r.get("total_profit", 0) + r.get("total_cost", 0)),
    }

@router.get(
    "/representatives/{name}/users",
    dependencies=[Depends(admin_required)]
)
async def representative_users_with_competition(
    name: str,
    year: int = Query(...),
    month: int = Query(...)
):
    rep_name = unquote(name)

    # 1️⃣ Yarışmayı bul (tüm statüler kabul, sadece yıla ve aya bakıyoruz)
    competition = await db.competitions.find_one({
        "year": year,
        "month": month,
    })

    # 2️⃣ Bu mümessilin tüm kullanıcılarını bul
    profiles_cursor = db.user_profiles.find({
        "$or": [
            {"representative": rep_name},
            {"representative.name": rep_name},
        ]
    })
    
    user_ids = []
    async for p in profiles_cursor:
        user_ids.append(p["user_id"])

    # 3️⃣ Katılanları belirle
    participant_ids = set()
    if competition:
        participants_cursor = db.competition_participants.find({
            "competition_id": competition["_id"],
            "user_id": {"$in": user_ids}
        })
        async for p in participants_cursor:
            participant_ids.add(p["user_id"])

    # 4️⃣ Kullanıcı detaylarını çek ve ayır
    users_cursor = db.users.find({ "_id": { "$in": user_ids } })
    
    in_competition = []
    not_in_competition = []

    async for u in users_cursor:
        user_data = {
            "id": str(u["_id"]),
            "name": u.get("full_name") or u.get("email") or "Bilinmeyen",
            "email": u.get("email"),
        }

        if u["_id"] in participant_ids:
            in_competition.append(user_data)
        else:
            not_in_competition.append(user_data)

    return {
        "competition": {
            "year": year,
            "month": month,
            "id": str(competition["_id"]) if competition else None,
            "status": competition.get("status") if competition else "none"
        },
        "in_competition": in_competition,
        "not_in_competition": not_in_competition,
    }


@router.get("/users/{user_id}", dependencies=[Depends(admin_required)])
async def admin_user_detail(user_id: str):
    if not ObjectId.is_valid(user_id):
        raise HTTPException(400, "Invalid user id")

    # 1️⃣ AGGREGATION (Lifetime Stats)
    pipeline = [
        { "$match": { "_id": ObjectId(user_id) } },

        {
            "$lookup": {
                "from": "user_profiles",
                "localField": "_id",
                "foreignField": "user_id",
                "as": "profile"
            }
        },
        { "$unwind": { "path": "$profile", "preserveNullAndEmptyArrays": True }},

        {
            "$lookup": {
                "from": "sales_reports",
                "localField": "_id",
                "foreignField": "user_id",
                "as": "reports"
            }
        },

        {
            "$addFields": {
                "total_reports": { "$size": "$reports" },
                "total_items": { "$sum": "$reports.summary.total_items" },
                "total_profit": { "$sum": "$reports.summary.total_profit" },
                "total_cost": { "$sum": "$reports.summary.total_cost" }
            }
        }
    ]

    doc = await db.users.aggregate(pipeline).to_list(1)
    if not doc:
        raise HTTPException(404, "User not found")

    u = doc[0]

    # 2️⃣ CONTEXT-AWARE COUNTS (Reset Logic)
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    current_month_count = await db.sales_reports.count_documents({
        "user_id": u["_id"],
        "createdAt": {"$gte": month_start},
        "is_competition_report": {"$ne": True}
    })

    # 3️⃣ COMPETITION INFO
    last_comp_info = None
    
    # En son katıldığı yarışmayı bul
    last_participation = await db.competition_participants.find_one(
        { "user_id": u["_id"] }, 
        sort=[("accepted_at", -1)]
    )

    if last_participation:
        comp = await db.competitions.find_one({"_id": last_participation["competition_id"]})
        if comp:
            # 🔥 GERÇEK STATUS KONTROLÜ (Tarih + Bireysel Bitiş)
            is_truly_active = (
                comp.get("status") == "active" and 
                comp["starts_at"] <= now <= comp["ends_at"] and
                (last_participation.get("finished_at") is None or last_participation["finished_at"] > now)
            )
            
            if not is_truly_active and last_participation.get("finished_at") and last_participation["finished_at"] <= now:
                # Kullanıcı bireysel olarak bitmişse, admin panelinde onu "Normal Mod"da (0/4 hedef) gösterelim.
                last_comp_info = None
            else:
                last_comp_info = {
                    "id": str(comp["_id"]),
                    "status": "active" if is_truly_active else comp.get("status"),
                    "year": comp["year"],
                    "month": comp["month"],
                    "is_finished_individually": last_participation.get("finished_at") is not None
                }

    return {
        "id": str(u["_id"]),
        "name": u.get("full_name"),
        "email": u.get("email"),
        "pharmacy": u.get("profile", {}).get("pharmacy_name"),
        "region": u.get("profile", {}).get("region"),
        "representative": u.get("profile", {}).get("representative"),
        
        # Lifetime Stats
        "total_reports": u.get("total_reports", 0),
        "total_items": u.get("total_items", 0),
        "total_profit": u.get("total_profit", 0),
        "total_revenue": u.get("total_profit", 0) + u.get("total_cost", 0),

        # Context Stats (for UI Reset & Download)
        "current_month_reports": current_month_count,
        "competition": last_comp_info
    }


@router.post("/users/{user_id}/notify", dependencies=[Depends(admin_required)])
async def send_user_notification(user_id: str, type: str = "report"):
    if not ObjectId.is_valid(user_id):
        raise HTTPException(400, "Invalid user id")

    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(404, "User not found")
        
    # CONTENT MAP
    content_map = {
        "report": {
            "title": "Rapor Hatırlatması 📊",
            "body": "Bu haftaki stok raporunu yüklemeyi unutma! Hemen şimdi yükle.",
        },
        "competition": {
            "title": "Yarışma Başladı! 🏆",
            "body": "Yeni yarışma kayıtları açıldı. Hemen katıl ve kazanmaya başla!",
        },
        "stock": {
            "title": "Stok Hatırlatması 📦",
            "body": "Lütfen stok durumunuzu mümessilinize bildiriniz.",
        }
    }
    
    content = content_map.get(type, content_map["report"])

    # 📌 SAVE TO DB & SEND PUSH (via service.py)
    from app.notifications.service import create_notification
    await create_notification(
        user_id=user["_id"],
        title=content["title"],
        body=content["body"],
        type=type
    )

    return {
        "message": "Bildirim gönderildi", 
        "status": "sent"
    }

@router.get("/competitions", dependencies=[Depends(admin_required)])
async def admin_list_competitions():
    cursor = db.competitions.find({}).sort("starts_at", -1)

    items = []
    async for c in cursor:
        items.append({
            "id": str(c["_id"]),
            "year": c["year"],
            "month": c["month"],
            "status": c.get("status", "upcoming"),
            "starts_at": c["starts_at"],
            "ends_at": c["ends_at"],
        })

    return items

@router.post("/competitions", dependencies=[Depends(admin_required)])
async def admin_create_competition(year: int, month: int):
    exists = await db.competitions.find_one({
        "year": year,
        "month": month,
        "status": { "$ne": "cancelled" }
    })

    if exists:
        raise HTTPException(400, "Competition already exists")

    starts_at = datetime(year, month, 1)
    ends_at = end_of_month_utc(year, month)

    doc = {
        "year": year,
        "month": month,
        "starts_at": starts_at,
        "ends_at": ends_at,
        "status": "upcoming",
        "created_at": datetime.utcnow(),
    }

    res = await db.competitions.insert_one(doc)

    return {
        "id": str(res.inserted_id),
        "status": "upcoming",
    }


from app.competitions.cron import activate_competition_by_admin

@router.post("/competitions/{competition_id}/start", dependencies=[Depends(admin_required)])
async def admin_start_competition(competition_id: str):
    if not ObjectId.is_valid(competition_id):
        raise HTTPException(400, "Invalid competition id")

    comp = await db.competitions.find_one({ "_id": ObjectId(competition_id) })
    if not comp:
        raise HTTPException(404, "Competition not found")

    if comp.get("status") != "upcoming":
        raise HTTPException(400, "Only upcoming competitions can be started")

    # 🛡️ ZAMAN KONTROLÜ
    now = datetime.utcnow()
    if comp["starts_at"] > now:
        raise HTTPException(400, "Henüz yarışma başlangıç tarihi gelmedi")

    # 🔒 AY BAZLI TEK YARIŞMA KURALI
    existing = await db.competitions.find_one({
        "_id": { "$ne": comp["_id"] },
        "year": comp["year"],
        "month": comp["month"],
        "status": "active",
    })

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Bu ay için zaten aktif bir yarışma var"
        )

    # 🔥 TEK SOURCE OF TRUTH
    await activate_competition_by_admin(comp["_id"])

    return { "status": "active" }




@router.post("/competitions/{competition_id}/finish", dependencies=[Depends(admin_required)])
async def admin_finish_competition(competition_id: str):
    comp = await db.competitions.find_one({ "_id": ObjectId(competition_id) })

    if not comp:
        raise HTTPException(404, "Competition not found")

    if comp.get("status") != "active":
        raise HTTPException(400, "Only active competitions can be finished")

    now = datetime.utcnow()

    # 1️⃣ Yarışmayı Bitir
    await db.competitions.update_one(
        { "_id": comp["_id"] },
        {
            "$set": {
                "status": "completed",
                "ended_at": now,
            }
        }
    )

    # 2️⃣ Katılımcıların da bittiği tarihi işle
    await db.competition_participants.update_many(
        { "competition_id": comp["_id"] },
        {
            "$set": {
                "finished_at": now
            }
        }
    )

    # 3️⃣ Bildirim Gönder (Mobilin Yenilenmesi İçin)
    from app.notifications.service import create_notification
    
    participants_cursor = db.competition_participants.find({ "competition_id": comp["_id"] })
    
    # Send individually or bulk? For simplicity and to use existing service, loop.
    # Ideally use a bulk service if available, but for now loop is acceptable for typical participant counts.
    async for p in participants_cursor:
        await create_notification(
            user_id=p["user_id"],
            title="Yarışma Sona Erdi",
            body=f"{comp['month']}/{comp['year']} yarışması tamamlanmıştır.",
            type="competition_ended",
            data={
                "competition_id": str(comp["_id"]),
                "refresh": "true"
            }
        )

    return { "status": "completed" }


@router.post("/competitions/{competition_id}/cancel", dependencies=[Depends(admin_required)])
async def admin_cancel_competition(competition_id: str):
    if not ObjectId.is_valid(competition_id):
        raise HTTPException(400, "Invalid competition id")

    comp = await db.competitions.find_one({ "_id": ObjectId(competition_id) })
    if not comp:
        raise HTTPException(404, "Competition not found")

    # ❌ SADECE COMPLETED KORUNUR
    if comp.get("status") == "completed":
        raise HTTPException(
            status_code=400,
            detail="Completed competition cannot be cancelled"
        )

    # 🔥 HARD DELETE (upcoming + active)
    await db.competitions.delete_one({ "_id": comp["_id"] })

    # 🔥 BAĞLI KAYITLARI TEMİZLE
    await db.competition_registrations.delete_many({
        "competition_id": comp["_id"]
    })

    await db.competition_participants.delete_many({
        "competition_id": comp["_id"]
    })

    return { "status": "deleted" }


@router.delete("/competitions/{competition_id}", dependencies=[Depends(admin_required)])
async def admin_delete_competition(competition_id: str):
    """Delete a completed competition and all related data."""
    if not ObjectId.is_valid(competition_id):
        raise HTTPException(400, "Invalid competition id")

    comp = await db.competitions.find_one({ "_id": ObjectId(competition_id) })
    if not comp:
        raise HTTPException(404, "Competition not found")

    # 🔥 HARD DELETE THE COMPETITION
    await db.competitions.delete_one({ "_id": comp["_id"] })

    # 🔥 BAĞLI KAYITLARI TEMİZLE
    await db.competition_registrations.delete_many({
        "competition_id": comp["_id"]
    })

    await db.competition_participants.delete_many({
        "competition_id": comp["_id"]
    })

    # 🔥 SCOREBOARD VERİLERİNİ TEMİZLE (varsa)
    await db.scoreboards.delete_many({
        "competition_id": comp["_id"]
    })

    return { "status": "deleted" }

@router.get(
    "/competitions/{competition_id}/participants",
    dependencies=[Depends(admin_required)]
)
async def admin_competition_participants(competition_id: str):
    if not ObjectId.is_valid(competition_id):
        raise HTTPException(400, "Invalid competition id")

    competition = await db.competitions.find_one({
        "_id": ObjectId(competition_id)
    })

    if not competition:
        raise HTTPException(404, "Competition not found")

    # 🔹 Yarışmaya katılan (kabul eden) kullanıcılar
    participants_cursor = db.competition_participants.find({
        "competition_id": competition["_id"]
    })
    
    # 🔹 Yarışmaya kayıtlı (henüz kabul etmemiş) kullanıcılar
    registrations_cursor = db.competition_registrations.find({
        "competition_id": competition["_id"]
    })

    user_status_map = {} # user_id_str -> status
    
    async for p in participants_cursor:
        user_status_map[str(p["user_id"])] = "accepted"
        
    async for r in registrations_cursor:
        # Eğer zaten accepted ise (üstte eklendiyse) ezme
        uid_str = str(r["user_id"])
        if uid_str not in user_status_map:
            user_status_map[uid_str] = "registered"

    user_ids = [ObjectId(uid) for uid in user_status_map.keys()]

    if not user_ids:
        return {
            "competition": {
                "year": competition["year"],
                "month": competition["month"],
            },
            "participants": []
        }

    print(f"📊 ADMIN COMP PARTICIPANTS | CompID: {competition_id} | Users to check: {len(user_ids)}")

    # 🔹 Kullanıcı + rapor sayısı
    pipeline = [
        { "$match": { "_id": { "$in": user_ids } } },

        {
            "$lookup": {
                "from": "sales_reports",
                "let": { "uid": "$_id" },
                "pipeline": [
                    { 
                        "$match": { 
                            "$expr": { 
                                "$and": [
                                    { "$eq": ["$user_id", "$$uid"] },
                                    { "$eq": ["$competition_id", ObjectId(competition_id)] },
                                    # 🔥 STRICT DATE FILTER:
                                    # Admin panelinde de sadece takvim aralığındakileri say.
                                    { "$gte": ["$createdAt", competition["starts_at"]] },
                                    { "$lte": ["$createdAt", competition["ends_at"]] }
                                ]
                            }
                        } 
                    }
                ],
                "as": "reports"
            }
        },

        {
            "$addFields": {
                "report_count": { "$size": "$reports" }
            }
        },

        {
            "$project": {
                "_id": 0,
                "id": { "$toString": "$_id" },
                "name": { "$ifNull": ["$full_name", "$email"] },
                "email": 1,
                "report_count": 1
            }
        }
    ]

    cursor = db.users.aggregate(pipeline)

    participants = []
    async for u in cursor:
        u["status"] = user_status_map.get(u["id"], "unknown")
        participants.append(u)
        
        if u["report_count"] > 0:
             print(f"🔴 DEBUG USER {u['name']}: {u['report_count']} reports found for comp {competition_id}")

    # Sıralama: Katıldı (accepted) önce, sonra rapor sayısı
    participants.sort(key=lambda x: (x["status"] != "accepted", -x["report_count"]))

    return {
        "competition": {
            "year": competition["year"],
            "month": competition["month"],
            "status": competition.get("status"),
        },
        "participants": participants
    }


# ================= FIELD VISITS (ADMIN) =================

@router.get("/field-visits/users", dependencies=[Depends(admin_required)])
async def admin_field_visit_users():
    """List all users who have field visits, with summary stats."""
    pipeline = [
        # Group by user_id
        {
            "$group": {
                "_id": "$user_id",
                "total_visits": {"$sum": 1},
                "confirmed_count": {
                    "$sum": {"$cond": [{"$eq": ["$confirmed", True]}, 1, 0]}
                },
                "evaluated_count": {
                    "$sum": {"$cond": [{"$ne": ["$evaluation", None]}, 1, 0]}
                },
                "last_visit_date": {"$max": "$visit_date"},
            }
        },
        # Join user
        {
            "$lookup": {
                "from": "users",
                "localField": "_id",
                "foreignField": "_id",
                "as": "user",
            }
        },
        {"$unwind": {"path": "$user", "preserveNullAndEmptyArrays": True}},
        # Join profile
        {
            "$lookup": {
                "from": "user_profiles",
                "localField": "_id",
                "foreignField": "user_id",
                "as": "profile",
            }
        },
        {"$unwind": {"path": "$profile", "preserveNullAndEmptyArrays": True}},
        # Sort by last visit
        {"$sort": {"last_visit_date": -1}},
    ]

    cursor = db.field_visits.aggregate(pipeline)
    result = []
    async for doc in cursor:
        user = doc.get("user", {})
        profile = doc.get("profile", {})

        # Normalize representative
        rep = profile.get("representative")
        if isinstance(rep, dict):
            rep = rep.get("name", "")

        result.append({
            "user_id": str(doc["_id"]),
            "full_name": user.get("full_name", "Bilinmeyen"),
            "email": user.get("email", ""),
            "pharmacy_name": profile.get("pharmacy_name", "-"),
            "district": profile.get("district", "-"),
            "region": profile.get("region", "-"),
            "representative": rep or "-",
            "total_visits": doc.get("total_visits", 0),
            "confirmed_count": doc.get("confirmed_count", 0),
            "evaluated_count": doc.get("evaluated_count", 0),
            "last_visit_date": doc.get("last_visit_date"),
        })

    return result


@router.get("/field-visits/user/{user_id}", dependencies=[Depends(admin_required)])
async def admin_field_visit_detail(user_id: str):
    """Get all field visits for a specific user."""
    if not ObjectId.is_valid(user_id):
        raise HTTPException(400, "Invalid user ID")

    oid = ObjectId(user_id)

    # Get user info
    user = await db.users.find_one({"_id": oid})
    if not user:
        raise HTTPException(404, "User not found")

    profile = await db.user_profiles.find_one({"user_id": oid})

    # Normalize representative
    rep = (profile or {}).get("representative")
    if isinstance(rep, dict):
        rep = rep.get("name", "")

    # Get visits
    cursor = db.field_visits.find({"user_id": oid}).sort("visit_date", -1)

    visits = []
    async for doc in cursor:
        evaluation = doc.get("evaluation")
        visits.append({
            "id": str(doc["_id"]),
            "visit_date": doc.get("visit_date"),
            "visit_time": doc.get("visit_time"),
            "pharmacy_name": doc.get("pharmacy_name", "-"),
            "pharmacy_district": doc.get("pharmacy_district", "-"),
            "notes": doc.get("notes", ""),
            "confirmed": doc.get("confirmed", False),
            "confirmed_at": doc.get("confirmed_at"),
            "evaluation": {
                "duration_hours": evaluation["duration_hours"],
                "transport_type": evaluation["transport_type"],
                "taxi_cost": evaluation.get("taxi_cost", 0),
                "pharmacist_rating": evaluation["pharmacist_rating"],
                "evaluation_notes": evaluation.get("evaluation_notes", ""),
            } if evaluation else None,
        })

    return {
        "user": {
            "id": str(user["_id"]),
            "full_name": user.get("full_name", "Bilinmeyen"),
            "email": user.get("email", ""),
            "pharmacy_name": (profile or {}).get("pharmacy_name", "-"),
            "district": (profile or {}).get("district", "-"),
            "region": (profile or {}).get("region", "-"),
            "representative": rep or "-",
        },
        "visits": visits,
    }


@router.get("/field-visits/export", dependencies=[Depends(admin_required)])
async def admin_field_visits_export(
    user_id: str = Query(None, description="Filter by user ID"),
):
    """Export field visits as Excel file."""
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse

    query = {}
    filename_suffix = "Tum_Kullanicilar"

    if user_id and ObjectId.is_valid(user_id):
        query["user_id"] = ObjectId(user_id)
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if user:
            safe_name = (user.get("full_name") or "kullanici").replace(" ", "_")
            filename_suffix = safe_name

    # Fetch visits
    cursor = db.field_visits.find(query).sort("visit_date", -1)
    visits = await cursor.to_list(None)

    # Build user map for names
    user_ids = list(set(v["user_id"] for v in visits))
    user_map = {}
    pharmacy_map = {}

    if user_ids:
        users = await db.users.find({"_id": {"$in": user_ids}}).to_list(None)
        for u in users:
            user_map[u["_id"]] = u.get("full_name", "Bilinmeyen")

        profiles = await db.user_profiles.find({"user_id": {"$in": user_ids}}).to_list(None)
        for p in profiles:
            pharmacy_map[p["user_id"]] = p.get("pharmacy_name", "-")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saha Ziyaretleri"

    # Headers
    headers = [
        "Kullanıcı", "Eczane", "İlçe", "Tarih", "Saat",
        "Durum", "Süre (Saat)", "Ulaşım", "Taksi Ücreti (₺)",
        "Eczacı Puanı", "Ziyaret Notları", "Değerlendirme Notları",
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    transport_labels = {
        "taksi": "Taksi",
        "kendi_araci": "Kendi Aracı",
        "toplu_tasima": "Toplu Taşıma",
        "yuruyus": "Yürüyüş",
    }

    for v in visits:
        uid = v.get("user_id")
        evaluation = v.get("evaluation")

        if v.get("evaluation"):
            status = "Değerlendirildi"
        elif v.get("confirmed"):
            status = "Gidildi"
        else:
            status = "Planlandı"

        row = [
            user_map.get(uid, "Bilinmeyen"),
            v.get("pharmacy_name", "-"),
            v.get("pharmacy_district", "-"),
            v.get("visit_date", "-"),
            v.get("visit_time", "-"),
            status,
            evaluation["duration_hours"] if evaluation else "",
            transport_labels.get(evaluation["transport_type"], evaluation["transport_type"]) if evaluation else "",
            evaluation.get("taxi_cost", 0) if evaluation and evaluation.get("transport_type") == "taksi" else "",
            evaluation["pharmacist_rating"] if evaluation else "",
            v.get("notes", ""),
            evaluation.get("evaluation_notes", "") if evaluation else "",
        ]
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Saha_Ziyareti_{filename_suffix}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )



